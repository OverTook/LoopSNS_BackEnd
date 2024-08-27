from app import db
import time
from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import requests
from firebase_admin import auth
from collections import Counter
import matplotlib.pyplot as plt
from matplotlib import rcParams

# 엑셀에 넣을 한글 폰트 깨져서 한글 폰트 설정
rcParams['font.family'] = 'NanumGothic'
rcParams['axes.unicode_minus'] = False  

# 마커 블루프린트 작성
report_routes = Blueprint('report', __name__)

# 보고서에서 다운로드 버튼 클릭시 다운로드 되는 부분
@report_routes.route('/report_download', methods=['GET'])
def report_download():
    intention = request.args.getlist('intentions')
    subject = request.args.getlist('subjects')

    sido = request.args.get('sido')
    sigungu = request.args.get('sigungu')
    eupmyeondong = request.args.get('eupmyeondong')

    articles = []
    articles_ref = db.collection('articles')
    
    # in 쿼리문을 사용하게 되면 최대 30개씩 필터링에 제한이 걸려 수정함.
    # 의도 리스트를 30개씩 잘라서 in 쿼리문을 진행 
    if intention:
        for i in range(0, len(intention), 30):
            cut_intention = intention[i:i + 30]
            query = articles_ref.where(field_path='intention', op_string='in', value=cut_intention)
            
            # 시도, 시군구, 읍면동 필터링
            if sido:
                query = query.where(field_path='sido', op_string='==', value=sido)
            if sigungu:
                query = query.where(field_path='sigungu', op_string='==', value=sigungu)
            if eupmyeondong:
                query = query.where(field_path='eupmyeondong', op_string='==', value=eupmyeondong)
                
            # 쿼리 실행
            articles_stream = query.stream()
            for article in articles_stream:
                article_dict = article.to_dict()
                              
                # subject 필터링
                if article_dict.get('subject') in subject:
                    user_id = article_dict.get('user_id')
                    user = auth.get_user(user_id)
                    user_nickname = user.display_name
                    
                    latlng = article_dict.get('latlng')  # 좌표
                    latitude = latlng.latitude
                    longitude = latlng.longitude
                    latlng = str(latitude) + ", "+ str(longitude)
                    
                    articles.append({
                        'intention': article_dict.get('intention'),
                        'subject': article_dict.get('subject'),
                        'contents': article_dict.get('contents', ''),
                        'name': user_nickname,
                        'time': article_dict.get('time', None).strftime("%Y-%m-%d %H:%M"),
                        'image_urls': article_dict.get('image_urls', []),
                        'latlng': latlng,
                        'address': article_dict.get('address'),
                        'sido': article_dict.get('sido'),
                        'sigungu': article_dict.get('sigungu'),
                        'eupmyeondong': article_dict.get('eupmyeondong')
                    })
                    
    # 엑셀 파일 생성
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Loop_report"

    # 헤더 작성
    headers = ['의도', '주제', '내용', '작성자', '시간', '이미지', '좌표', '상세주소', '시/도', '시/군/구', '읍/면/동']
    worksheet.append(headers)

    # 데이터 추가
    for row_idx, article in enumerate(articles, start=2):  # 헤더가 첫 줄이므로 2번째 줄부터 시작
        worksheet.append([
            article['intention'],
            article['subject'],
            article['contents'],
            article['name'],
            article['time'],
            '',  # 이미지 칸 (나중에 이미지 삽입할 곳임)
            article['latlng'],
            article['address'],
            article['sido'],
            article['sigungu'],
            article['eupmyeondong']
        ])

        # 이미지 삽입
        for idx, url in enumerate(article['image_urls']):
            response = requests.get(url)
            img = OpenpyxlImage(BytesIO(response.content))
            img.width, img.height = 100, 100  # 이미지 크기 조정
            col = get_column_letter(6 + idx)  # 이미지가 삽입될 열
            worksheet.add_image(img, f"{col}{row_idx}")
            
            # 이미지 크기에 맞춰서 셀 크기 조정
            worksheet.column_dimensions[col].width = 15  # 열 너비를 적절히 조정
            worksheet.row_dimensions[row_idx].height = 100  # 행 높이를 이미지 높이에 맞춤

    # Excel 워크시트에서 자동 필터를 추가하는 부분 (필터 넣으려면 중요한 부분!)
    worksheet.auto_filter.ref = worksheet.dimensions

    # 의도와 주제 데이터 갯수 세기
    intentions_count = Counter([article['intention'] for article in articles])
    subjects_count = Counter([article['subject'] for article in articles])

    # 막대 그래프를 만들고 이미지화하여 엑셀에 삽입하는 함수
    def bar_chart(data, title):
        # 막대 그래프에 색상을 결정하기 위한 리스트
        colors = ['#17becf', '#98df8a', '#ff9896', '#c5b0d5', '#ffbb78', '#f7b7a3']
        
        plt.figure(figsize=(10, 5))
        
        keys = list(data.keys())
        values = list(data.values())
        
        # 막대에 색상을 인덱스에 따라서 설정해줌
        bar_colors = [colors[i % len(colors)] for i in range(len(keys))]
        
        plt.bar(keys, values, color=bar_colors)
        
        plt.title(title)
        plt.ylabel('갯수')

        # 그래프를 BytesIO 객체에 저장해서 반환하도록 함
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        plt.close()
        return img_data

    # 원형 그래프를 만들고 이미지화하여 엑셀에 삽입하는 함수
    def pie_chart(data, title):
        # 원형 그래프에 색상을 결정하기 위한 리스트
        colors = ['#ff9999','#66b3ff','#99ff99','#ffcc99','#c2c2f0','#ffb3e6']
        
        # explode 설정: 각 조각을 조금씩 분리
        explode = [0.05] * len(data)
        
        plt.figure(figsize=(8, 6))
        
        labels = list(data.keys())
        sizes = list(data.values())
        
        plt.pie(sizes, labels=labels, colors=colors, explode=explode, autopct='%1.1f%%', startangle=140)
        
        plt.title(title)
        
        # 그래프를 BytesIO 객체에 저장해서 반환하도록 함
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        plt.close()
        return img_data

    # 의도 그래프
    intention_chart = bar_chart(intentions_count, "의도 그래프")
    intention_img = OpenpyxlImage(intention_chart)
    intention_img.width, intention_img.height = 640, 480 # 크기 설정
    worksheet.add_image(intention_img, f'N{2}') # A(n+3)번 칸에 이미지 삽입 f'A{len(articles) + 3}'

    # 주제 그래프
    subject_chart = bar_chart(subjects_count, "주제 그래프")
    subject_img = OpenpyxlImage(subject_chart)
    subject_img.width, subject_img.height = 640, 480 # 크기 설정
    worksheet.add_image(subject_img, f'N{24}') # A(n+25)번 칸에 이미지 삽입

    # 의도 원형 그래프
    intention_pie_chart = pie_chart(intentions_count, "의도 원형 그래프")
    intention_pie_img = OpenpyxlImage(intention_pie_chart)
    intention_pie_img.width, intention_pie_img.height = 640, 480  # 크기 설정
    worksheet.add_image(intention_pie_img, f'X{2}')  # A(n+35)번 칸에 이미지 삽입

    # 주제 원형 그래프
    subject_pie_chart = pie_chart(subjects_count, "주제 원형 그래프")
    subject_pie_img = OpenpyxlImage(subject_pie_chart)
    subject_pie_img.width, subject_pie_img.height = 640, 480  # 크기 설정
    worksheet.add_image(subject_pie_img, f'X{24}')  # A(n+55)번 칸에 이미지 삽입

    # 엑셀 파일을 메모리에 저장
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # 엑셀 파일을 프론트로 전송
    filename = f"report_{int(time.time())}.xlsx"

    time.sleep(5)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )